import openpyxl
import pprint

wb = openpyxl.load_workbook('censuspopdata.xlsx')
ws = wb.active

countyPopData = {}
# CensusTract State County POP2010
# for row in ws[1]:
#     print(row.value, end=" ")

# {'AL': {'Autauga': {'pop': 54571, 'tracts': 12}}
# CensusTract	State	County	POP2010
# 01001020100	AL	Autauga	1912
for i in range(2, ws.max_row+1):
    state = ws.cell(row=i, column=2).value
    county = ws.cell(row=i, column=3).value
    pop = ws.cell(row=i, column=4).value

    countyPopData.setdefault(state, {})
    countyPopData[state].setdefault(county, {'tracts': 0, 'pop': 0})

    countyPopData[state][county]['tracts'] += 1
    countyPopData[state][county]['pop'] += int(pop)

resultFile = open('censusResult.py', 'w')
resultFile.write('censusResult = ' + pprint.pformat(countyPopData))
resultFile.close()