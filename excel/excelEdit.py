import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook('example.xlsx')
print(wb.sheetnames)

for sheet in wb:
    print(sheet.title)

mySheet = wb.create_sheet('mySheet')
print(wb.sheetnames)

sheet3 = wb.get_sheet_by_name('Sheet3')
print(sheet3)

sheet4 = wb['mySheet']
print(sheet4)

# - active 活动的sheet页
ws = wb.active
print(ws)
print(ws['A1'])
print(ws['A1'].value)

# - coordinate 直接显示单元格位置
c = ws['B1']
print('Row {} Colum {} is {}'.format(c.row, c.column, c.value))
print('Cell {} is {}'.format(c.coordinate, c.value))

print(ws.cell(row=2, column=2))
print(ws.cell(row=2, column=2).value)

for i in range(1, 8, 2):
    print(ws.cell(row=i, column=2).value)

# - 获取列元素
colC = ws['C']
print(colC)
for cell in colC:
    print(cell.value)

# - 获取行元素
row6 = ws[6]
print(row6)    
for row in row6:
    print(row.value)

# - 切片，遍历两列数据
col_range = ws['B:C']
for col in col_range:
    for cell in col:
        print(cell.value)   

# - 切片，遍历多行数据
row_range = ws[2:6]
for row in row_range:
    for cell in row:
        print(cell.value) 
print('------------------------------------------------')
# - 迭代，行与列
for row in ws.iter_rows(min_row=1, max_row=4, max_col=2):
    for cell in row:
        print(cell.value)

# - 按行遍历
print('------------------------------------------------')
print(tuple(ws.rows))

# - 按列遍历
print('------------------------------------------------')
print(tuple(ws.columns))

print('------------------------------------------------')
cell_range = ws['A1':'C3']
for cell in cell_range:
    for c in cell:
        print(c.coordinate, c.value)

print('Total Rows is {}, Total Columns is {}'.format(ws.max_row, ws.max_column))

print(get_column_letter(2), get_column_letter(20), get_column_letter(32))
print(column_index_from_string('ABQ'))